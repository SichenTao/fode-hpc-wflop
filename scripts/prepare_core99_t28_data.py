#!/usr/bin/env python3
"""
WFLOP IMPLEMENTATION FACT DECLARATION
Implementation unit: T28 author-data extraction utility
Paper DOI: 10.5194/wes-10-1661-2025
Public source: https://github.com/Twitwi96/WINDFLOWER/tree/v1.0.0
Public data DOI: 10.5281/zenodo.13946931
Source revision: 427ee84dfdb9fb275b229e8eecf4503071d8fed4
License: AGPL-3.0.
Missing/conflicting facts: the paper says 216 MW while the public turbine
curve and 72 * 3.075 MW imply 221.4 MW; the implementation follows the
public code and curve (221.4 MW), as later paper tables and figures do.
Resolution: this script performs a lossless numeric extraction of the
paper-linked XLSX assets; it does not invent or interpolate source rows.
Claim boundary: academic reconstruction from author-released data, not the
author software.
Last evidence-audit date: 2026-07-31
END WFLOP IMPLEMENTATION FACT DECLARATION
"""
from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path

from openpyxl import load_workbook


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1 << 20), b""):
            digest.update(block)
    return digest.hexdigest()


def number(value: object) -> str:
    if value is None or isinstance(value, str):
        return "nan"
    return format(float(value), ".17g")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    arguments = parser.parse_args()
    arguments.output.mkdir(parents=True, exist_ok=True)

    northwind = arguments.source / "Northwind.xlsx"
    scenarios = arguments.source / "data_scenarios.xlsx"
    turbine_book = load_workbook(northwind, read_only=True, data_only=True)

    layout_sheet = turbine_book["WT_coord"]
    layout_rows = list(layout_sheet.iter_rows(min_row=2, values_only=True))
    (arguments.output / "northwind_layout.tsv").write_text(
        "x_m\ty_m\tedge_flag\tedge_x_m\tedge_y_m\n"
        + "".join(
            "\t".join(
                (
                    number(row[4]),
                    number(row[5]),
                    "1" if row[6] else "0",
                    number(row[8]),
                    number(row[9]),
                )
            )
            + "\n"
            for row in layout_rows
        ),
        encoding="utf-8",
    )

    curve_sheet = turbine_book["V112-3"]
    curve_rows = list(curve_sheet.iter_rows(min_row=2, max_col=4, values_only=True))
    curve_rows = [row for row in curve_rows if row[0] is not None]
    (arguments.output / "v112_3_power_ct.tsv").write_text(
        "wind_speed_mps\tpower_mw\tcp\tct\n"
        + "".join("\t".join(number(value) for value in row) + "\n" for row in curve_rows),
        encoding="utf-8",
    )

    scenario_book = load_workbook(scenarios, read_only=True, data_only=True)
    counts: dict[str, int] = {}
    for year in ("2023", "2024"):
        sheet = scenario_book[year]
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        counts[year] = len(rows)
        (arguments.output / f"belgium_{year}.tsv").write_text(
            "wind_speed_mps\twind_direction_deg\tprice_da_eur_mwh\t"
            "price_capacity_up_eur_mw_h\tprice_activation_up_eur_mwh\t"
            "volume_activation_up_pu\tprice_imbalance_eur_mwh\n"
            + "".join(
                "\t".join(number(row[index]) for index in (4, 5, 6, 7, 8, 10, 15))
                + "\n"
                for row in rows
            ),
            encoding="utf-8",
        )

    receipt = {
        "schema_version": 1,
        "corpus_id": "T28",
        "paper_doi": "10.5194/wes-10-1661-2025",
        "data_doi": "10.5281/zenodo.13946931",
        "source_revision": "427ee84dfdb9fb275b229e8eecf4503071d8fed4",
        "license": "AGPL-3.0",
        "inputs": {
            northwind.name: sha256(northwind),
            scenarios.name: sha256(scenarios),
        },
        "scenario_rows": counts,
        "outputs": {},
    }
    for path in sorted(arguments.output.glob("*.tsv")):
        receipt["outputs"][path.name] = {
            "sha256": sha256(path),
            "bytes": path.stat().st_size,
        }
    (arguments.output / "extraction_receipt.json").write_text(
        json.dumps(receipt, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
